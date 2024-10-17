
from django import forms
from django.forms import RadioSelect, Select
from investments.models import Attachment, Package, Investment
from .models import AdministrativeLevel, Project, Sector
from django.core.exceptions import NON_FIELD_ERRORS
from django.utils.translation import gettext_lazy as _
from openpyxl import load_workbook
from django.core.exceptions import ValidationError


class UpdateInvestmentForm(forms.ModelForm):

    class Meta:
        model = Investment
        fields = ['physical_execution_rate', 'project_status', 'latitude', 'longitude']


class ProjectForm(forms.ModelForm):

    start_date = forms.DateField(input_formats=["%d/%m/%Y"])
    end_date = forms.DateField(input_formats=["%d/%m/%Y"])

    class Meta:
        model = Project
        exclude = ['id', 'created_date', 'updated_date', 'owner']
        labels = {
            'name': _("Name"),
            'description': _("Description"),
            'organization': _("Organization"),
            'start_date': _("Start Date"),
            'end_date': _("End Date"),
            'total_amount': _("Total Amount"),
            'sector': _("Sector"),
        }

    def __init__(self, *args, **kwargs):
        self.owner = kwargs.pop('owner', None)
        super().__init__(*args, **kwargs)

    def save(self, commit=True):
        instance = super().save(commit=False)
        if self.owner is not None:
            instance.owner = self.owner
            instance.organization = self.owner.organization
        return super().save(commit=commit)


class BulkUploadInvestmentsForm(forms.Form):
    xlsx_file = forms.FileField(label="XLSX File")
    headers = [
        'village_id',
        'ranking',
        'title',
        'sector_id',
        'estimated_cost',
        'start_date',
        'duration',
        'physical_execution_rate',
        'financial_implementation_rate',
        'project_status',
    ]

    def __init__(self, *args, **kwargs):
        self.project = kwargs.pop('project', None)
        super().__init__(*args, **kwargs)

    def clean_xlsx_file(self):
        """Validate if the uploaded file is a valid XLSX file."""
        uploaded_file = self.cleaned_data['xlsx_file']
        try:
            # Load the XLSX file
            wb = load_workbook(uploaded_file)
            sheet = wb.active
            headers = [cell.value for cell in sheet[1]]  # Read the first row for headers
        except Exception as e:
            raise ValidationError(f"Invalid XLSX file: {e}")
        return uploaded_file

    def save(self):
        project = self.project
        package = project.packages.all().first() or Package.objects.create(
            project=project,
            user=project.owner
        )
        investments = []
        uploaded_file = self.cleaned_data['xlsx_file']

        # Load the XLSX file
        wb = load_workbook(uploaded_file)
        sheet = wb.active  # Use the first sheet

        headers_dict = {}
        init_headers = list(self.headers)  # Make a copy of the headers

        for idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            if idx == 1:
                # Read the headers
                for k, header in enumerate(row):
                    normalized_header = header.strip().lower() if header else None
                    if normalized_header in init_headers:
                        headers_dict[normalized_header] = k
                        init_headers.remove(normalized_header)

                if init_headers:
                    raise ValidationError(f"Missing required headers: {', '.join(init_headers)}")

            else:
                try:
                    investments.append(Investment(
                        title=row[headers_dict['title']],
                        administrative_level=AdministrativeLevel.objects.get(id=row[headers_dict['village_id']]),
                        sector=Sector.objects.get(id=row[headers_dict['sector_id']]),
                        estimated_cost=row[headers_dict['estimated_cost']],
                        start_date=row[headers_dict['start_date']],
                        duration=row[headers_dict['duration']],
                        physical_execution_rate=row[headers_dict['physical_execution_rate']],
                        financial_implementation_rate=row[headers_dict['financial_implementation_rate']],
                        project_status=row[headers_dict['project_status']],
                        delays_consumed=0
                    ))
                except (KeyError, IndexError, ValueError) as e:
                    raise ValidationError(f"Error processing row {idx}: {e}")

        # Bulk create all investments
        Investment.objects.bulk_create(investments)

        # Add the investments to the package's funded investments
        package.funded_investments.add(*investments)

        return project


class AdministrativeLevelForm(forms.ModelForm):
    def __init__(self, parent: str = None, type: str = None, *args, **kwargs):
        super(AdministrativeLevelForm, self).__init__(*args, **kwargs)
        for label, field in self.fields.items():
            self.fields[label].widget.attrs.update({"class": "form-control"})
            if label == "parent":
                print("******* ", parent)
                self.fields[label].queryset = AdministrativeLevel.objects.filter(
                    type=parent
                )
                self.fields[label].label = parent
            if label == "type":
                self.fields[label].initial = type
                print("######### ", parent)

    class Meta:
        model = AdministrativeLevel
        exclude = [
            "no_sql_db_id",
            "geographical_unit",
            "cvd",
            "latitude",
            "longitude",
            "frontalier",
            "rural",
        ]  # specify the fields to be hid


# SearchVillages
class VillageSearchForm(forms.Form):
    region = forms.ModelChoiceField(
        queryset=AdministrativeLevel.objects.filter(type="Region"),
        required=False,
        empty_label=_("Toutes les régions"),
        label=_("Region"),
    )
    prefecture = forms.ModelChoiceField(
        queryset=AdministrativeLevel.objects.filter(type="Prefecture"),
        required=False,
        label=_("Prefecture"),
    )
    commune = forms.ModelChoiceField(
        queryset=AdministrativeLevel.objects.filter(type="Commune"),
        required=False,
        label=_("Commune"),
    )
    canton = forms.ModelChoiceField(
        queryset=AdministrativeLevel.objects.filter(type="Canton"),
        required=False,
        label=_("Canton"),
    )

class FinancialPartnerForm(forms.Form):
    name = forms.CharField()
    is_full_contribution = forms.BooleanField()
    potential_date = forms.DateField()
    commentaries = forms.CharField(widget=forms.Textarea())


class AttachmentFilterForm(forms.Form):
    TYPE_CHOICES = (
        (Attachment.PHOTO, _("Photo")),
        (Attachment.DOCUMENT, _("Document")),
        (None, _("Both")),
    )
    type = forms.ChoiceField(choices=TYPE_CHOICES, widget=RadioSelect, label=_("Type"))
    phase = forms.ChoiceField(widget=Select(attrs={'empty-option': '---------'}), required=False, label=_("Phase"))
    activity = forms.ChoiceField(widget=Select(attrs={'empty-option': '---------'}), required=False, label=_("Activity"))
    task = forms.ChoiceField(widget=Select(attrs={'empty-option': '---------'}), required=False, label=_("Task"))
    administrative_level = forms.ModelChoiceField(
        queryset=AdministrativeLevel.objects.filter(type=AdministrativeLevel.VILLAGE),
        widget=Select(attrs={'empty-option': '---------'}), required=False, label=_("Administrative level")
    )
